"""Strict Excel (.xlsx) import for Shoot quiz sets."""

from __future__ import annotations

import io
from typing import Any

from django.db import transaction

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

from .models import Question, QuizSet
from .quiz_set_utils import add_question_to_quiz_set

MAX_IMPORT_ROWS = 100
MAX_FILE_BYTES = 2 * 1024 * 1024

REQUIRED_HEADERS = [
    '题型', '题目', '选项A', '选项B', '选项C', '选项D', '正确答案', '时限秒',
]

HEADER_TO_FIELD = {
    '题型': 'question_type',
    '题目': 'text',
    '选项A': 'option_a',
    '选项B': 'option_b',
    '选项C': 'option_c',
    '选项D': 'option_d',
    '正确答案': 'correct_option',
    '时限秒': 'time_limit',
}

TYPE_ALIASES = {
    'single': Question.TYPE_SINGLE,
    '单选': Question.TYPE_SINGLE,
    '单选题': Question.TYPE_SINGLE,
    'multiple': Question.TYPE_MULTIPLE,
    '多选': Question.TYPE_MULTIPLE,
    '多选题': Question.TYPE_MULTIPLE,
    'judgment': Question.TYPE_JUDGMENT,
    'judgement': Question.TYPE_JUDGMENT,
    '判断': Question.TYPE_JUDGMENT,
    '判断题': Question.TYPE_JUDGMENT,
    'short_answer': Question.TYPE_SHORT_ANSWER,
    'shortanswer': Question.TYPE_SHORT_ANSWER,
    '简答': Question.TYPE_SHORT_ANSWER,
    '简答题': Question.TYPE_SHORT_ANSWER,
    'word_cloud': Question.TYPE_WORD_CLOUD,
    'wordcloud': Question.TYPE_WORD_CLOUD,
    '词云': Question.TYPE_WORD_CLOUD,
    '词云题': Question.TYPE_WORD_CLOUD,
    'explanation': Question.TYPE_EXPLANATION,
    '解释': Question.TYPE_EXPLANATION,
    '解释题': Question.TYPE_EXPLANATION,
}

EXAMPLE_ROWS = [
    {
        '题型': '单选题',
        '题目': '中国的首都是？',
        '选项A': '北京',
        '选项B': '上海',
        '选项C': '广州',
        '选项D': '深圳',
        '正确答案': 'A',
        '时限秒': 20,
    },
    {
        '题型': '多选题',
        '题目': '下列哪些是偶数？',
        '选项A': '2',
        '选项B': '3',
        '选项C': '4',
        '选项D': '5',
        '正确答案': 'A,C',
        '时限秒': 30,
    },
    {
        '题型': '判断题',
        '题目': '地球是圆的。',
        '选项A': '正确',
        '选项B': '错误',
        '选项C': '',
        '选项D': '',
        '正确答案': 'A',
        '时限秒': 15,
    },
    {
        '题型': '简答题',
        '题目': '水的化学式？',
        '选项A': 'H2O',
        '选项B': '',
        '选项C': '',
        '选项D': '',
        '正确答案': 'A',
        '时限秒': 25,
    },
]


class ExcelImportError(Exception):
    def __init__(self, message: str, row: int | None = None):
        self.row = row
        super().__init__(message)


def _cell_str(value: Any) -> str:
    if value is None:
        return ''
    if isinstance(value, bool):
        return '1' if value else '0'
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip()


def _normalize_type(raw: str) -> str:
    key = raw.strip().lower()
    if key in TYPE_ALIASES:
        return TYPE_ALIASES[key]
    if raw.strip() in TYPE_ALIASES:
        return TYPE_ALIASES[raw.strip()]
    raise ExcelImportError(
        f'题型无效「{raw}」，须为：单选题 / 多选题 / 判断题 / 简答题（或 single / multiple / judgment / short_answer）。解释题请在可视化编辑器中上传图片。'
    )


def _parse_correct_keys(raw: str) -> list[str]:
    text = raw.strip().upper().replace('，', ',').replace(' ', '')
    if not text:
        return []
    if ',' in text:
        return sorted({p for p in text.split(',') if p in ('A', 'B', 'C', 'D')})
    if text in ('A', 'B', 'C', 'D'):
        return [text]
    return sorted({c for c in text if c in ('A', 'B', 'C', 'D')})


def validate_question_row(row_num: int, raw: dict[str, str]) -> dict[str, Any]:
    qtype_raw = raw.get('question_type', '')
    text = raw.get('text', '')
    if not qtype_raw:
        raise ExcelImportError('题型不能为空', row_num)
    if not text:
        raise ExcelImportError('题目不能为空', row_num)

    question_type = _normalize_type(qtype_raw)
    option_a = raw.get('option_a', '')
    option_b = raw.get('option_b', '')
    option_c = raw.get('option_c', '')
    option_d = raw.get('option_d', '')
    correct_raw = raw.get('correct_option', '')
    time_raw = raw.get('time_limit', '20')

    try:
        time_limit = int(float(time_raw or 20))
    except (TypeError, ValueError):
        raise ExcelImportError(f'时限秒无效「{time_raw}」', row_num)
    time_limit = max(5, min(120, time_limit))

    if question_type == Question.TYPE_SHORT_ANSWER:
        if not option_a:
            raise ExcelImportError('简答题须在「选项A」列填写参考答案（多个用 | 分隔）', row_num)
        option_b = Question.TEXT_OPTION_PLACEHOLDER
        option_c = Question.TEXT_OPTION_PLACEHOLDER
        option_d = Question.TEXT_OPTION_PLACEHOLDER
        correct_option = 'A'
    elif question_type == Question.TYPE_WORD_CLOUD:
        option_a = Question.TEXT_OPTION_PLACEHOLDER
        option_b = Question.TEXT_OPTION_PLACEHOLDER
        option_c = Question.TEXT_OPTION_PLACEHOLDER
        option_d = Question.TEXT_OPTION_PLACEHOLDER
        correct_option = ''
    elif question_type == Question.TYPE_EXPLANATION:
        raise ExcelImportError('解释题只能在可视化编辑器中上传一张图片，无法通过 Excel 导入', row_num)
    elif question_type == Question.TYPE_JUDGMENT:
        option_a = option_a or '正确'
        option_b = option_b or '错误'
        option_c = Question.JUDGMENT_OPTION_PLACEHOLDER
        option_d = Question.JUDGMENT_OPTION_PLACEHOLDER
        keys = _parse_correct_keys(correct_raw)
        if len(keys) != 1 or keys[0] not in ('A', 'B'):
            raise ExcelImportError('判断题正确答案须为 A 或 B', row_num)
        correct_option = keys[0]
    elif question_type == Question.TYPE_MULTIPLE:
        if not all([option_a, option_b, option_c, option_d]):
            raise ExcelImportError('多选题四个选项均不能为空', row_num)
        keys = _parse_correct_keys(correct_raw)
        if len(keys) < 2:
            raise ExcelImportError('多选题正确答案须至少 2 个字母，逗号分隔，如 A,C', row_num)
        correct_option = ','.join(keys)
    else:
        if not all([option_a, option_b, option_c, option_d]):
            raise ExcelImportError('单选题四个选项均不能为空', row_num)
        keys = _parse_correct_keys(correct_raw)
        if len(keys) != 1:
            raise ExcelImportError('单选题正确答案须为单个字母 A/B/C/D', row_num)
        correct_option = keys[0]

    return {
        'text': text[:500],
        'question_type': question_type,
        'option_a': option_a[:200],
        'option_b': option_b[:200],
        'option_c': option_c[:200],
        'option_d': option_d[:200],
        'correct_option': correct_option[:10],
        'time_limit': time_limit,
    }


def read_rows_from_xlsx(file_bytes: bytes) -> list[dict[str, str]]:
    if len(file_bytes) > MAX_FILE_BYTES:
        raise ExcelImportError(f'文件超过 {MAX_FILE_BYTES // 1024 // 1024}MB 限制')

    try:
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise ExcelImportError(f'无法读取 Excel 文件：{exc}')

    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise ExcelImportError('表格为空')

    headers = [_cell_str(h) for h in header_row]
    if headers != REQUIRED_HEADERS:
        raise ExcelImportError(
            '表头必须严格为：'
            + ' | '.join(REQUIRED_HEADERS)
        )

    parsed_rows: list[dict[str, str]] = []
    excel_row_num = 2
    for row_values in rows_iter:
        cells = list(row_values or [])
        if not any(c is not None and _cell_str(c) for c in cells):
            excel_row_num += 1
            continue

        if len(cells) < len(REQUIRED_HEADERS):
            cells.extend([None] * (len(REQUIRED_HEADERS) - len(cells)))

        row_dict: dict[str, str] = {}
        for idx, header in enumerate(REQUIRED_HEADERS):
            field = HEADER_TO_FIELD[header]
            row_dict[field] = _cell_str(cells[idx])

        if not row_dict['text']:
            excel_row_num += 1
            continue

        parsed_rows.append({'excel_row': excel_row_num, **row_dict})
        excel_row_num += 1

        if len(parsed_rows) > MAX_IMPORT_ROWS:
            raise ExcelImportError(f'最多导入 {MAX_IMPORT_ROWS} 道题')

    workbook.close()

    if not parsed_rows:
        raise ExcelImportError('未找到有效题目行（表头下至少需一行数据）')

    return parsed_rows


def build_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = '题目'
    ws.append(REQUIRED_HEADERS)
    for row in EXAMPLE_ROWS:
        ws.append([row[h] for h in REQUIRED_HEADERS])

    col_widths = {
        'A': 14,
        'B': 30,
        'C': 18,
        'D': 18,
        'E': 18,
        'F': 18,
        'G': 14,
        'H': 12,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    dv = DataValidation(
        type='list',
        formula1='"单选题,多选题,判断题,简答题"',
        allow_blank=True,
    )
    dv.error = '请从下拉列表中选择有效题型（单选题、多选题、判断题、简答题）'
    dv.errorTitle = '题型无效'
    dv.prompt = '请点击右侧下拉箭头选择题型'
    dv.promptTitle = '选择题型'
    ws.add_data_validation(dv)
    dv.add('A2:A200')

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@transaction.atomic
def import_quiz_set_from_xlsx(teacher, title: str, file_bytes: bytes) -> QuizSet:
    title = (title or '').strip()
    if not title:
        raise ExcelImportError('请填写 Shoot 名称')

    raw_rows = read_rows_from_xlsx(file_bytes)
    validated = []
    for item in raw_rows:
        row_num = item['excel_row']
        data = validate_question_row(
            row_num,
            {k: v for k, v in item.items() if k != 'excel_row'},
        )
        validated.append(data)

    quiz_set = QuizSet.objects.create(title=title[:200], teacher=teacher, is_public=False)
    for order, item in enumerate(validated):
        question = Question.objects.create(
            teacher=teacher,
            is_public=False,
            **item,
        )
        add_question_to_quiz_set(quiz_set, question, order=order)

    return quiz_set
